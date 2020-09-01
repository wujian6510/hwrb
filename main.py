# 需求：
# 原始数据 2G小区表.xlsx [网元编号	小区网管标识	小区网管名称	省份	城市	区县	乡镇	所属城市区域	所属行政区域类型	是否采集MR	采集点标识	设备厂家CELL标识	小区别名	所属BTS标识	设备厂家	位置区码LAC	小区CI	频点类型	配置载频数	可用载频数	SDCCH配置信道数	静态配置PDCH信道数	SDCCH可用数	TCH配置信道数	小区BCCH频点	小区BSIC	小区TCH频点	运行状态	经度	纬度	载频最大发射功率	小区覆盖类型	相邻小区列表	载波类型	跳频模式	小区RA	是否开通半速率	GPRS开通情况	EDGE开通情况	增强全速率开通情况	所属MSC标识	所属BSC标识	关联物理站址编号	天线挂高	天线方向角	电子下倾角	机械倾角	自定义字段1	自定义字段2	自定义字段3	自定义字段4	自定义字段5	自定义字段6	自定义字段7	自定义字段8	自定义字段9	自定义字段10	DN	关联的机房/设备放置点编号	经度（WG84）	纬度（WG84）]
# 原始数据 2G-0729.xlsx [时间  小区名称	LAC	CI	所属BSC/RNC	网管运行状态	行政区名称	TCH话务量(erl)	GPRS上行LLC层吞吐率	GPRS下行LLC层吞吐率	TCH每线话务量(erl/信道)]
# 原始数据 3G-0729.xlsx [时间	LAC	CI	小区名	NodeB名称	RNC名称	CS域话务量  (含切)[erl]	RLC层上行业务流量(KByte)[KByte]	RLC层下行业务流量(KByte)[KByte]	小区载频平均接收功率[dBm]	RRC建立成功次数[次]	无线接通率[%]	CS域RAB异常释放的次数[次]	CS域RAB释放的次数[次]	CS域掉话率[%]	PS域掉线率[%]	PS域RAB异常释放的次数[次]	软切换成功率[%]	软切换成功次数[次]	同频硬切换出成功率[%]	RRC建立成功率[%]	RAB指配建立成功率[%]	异频硬切换出成功率[%]	异频硬切换出成功次数[次]]
# 原始数据 4G-0729.xlsl [时间	小区名称	CELL_ID	ENODEB_ID	TAC	网管运行状态	行政区名称	空口上行业务流量	空口下行业务流量	下行PRB平均利用率	RRC连接平均数	平均每PRB干扰噪声功率	小区级下行单用户平均感知速率]

# 目标数据
# sheet 234G话务流量汇总  [START_TIME	2G话务量	2G流量GB	3G话务量	3G流量GB	4G流量GB]
# sheet 物理基站流量汇总 [所属县分	物理站址名称	网络制式	归属场景	归属乡镇	4G流量（GB)	3G流量（GB)	3/4G总流量]
# sheet 4G话务流量 [时间	物理站点名称	行政区名称	小区名称	CELL_ID	ENODEB_ID	TAC	ENODEB_ID+CELL_ID	总流量(GB)	下行PRB平均利用率	RRC连接平均数	平均每PRB干扰噪声功率	小区级下行单用户平均感知速率	RRU型号	小区下行系统频域带宽	频段指示]
# sheet 3G话务流量 [时间 LAC	CI	LAC+CI	物理基站名称	行政区名称	小区名	NodeB名称	RNC名称	话务量	总流量（GB）	小区载频平均接收功率[dBm]	是否3Gonly小区]
# sheet 2G话务流量 [时间	物理站点名称	行政区名称	小区名称	LAC	CI	LAC+CI	所属BSC/RNC	TCH话务量(erl)	总流量]

import openpyxl
from distutils.core import setup
import py2exe

setup(console=["main.py"])

# 获取2G小区物理基站基础数据,存起来后面会用到
G2R_wb = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/2G小区表.xlsx')
G2R_sheet = G2R_wb.active
G2R = {}
for row in range(2, G2R_sheet.max_row + 1):
   ci = G2R_sheet['Q' + str(row)].value
   wljz = G2R_sheet['BA' + str(row)].value
   G2R[ci] = wljz
print(G2R)
# 写入Excel数据
wbRes = openpyxl.Workbook()
wbRes_sheet234G = wbRes.create_sheet(index=0,  title='234G话务流量汇总')
wbRes_sheet234G.append(['START_TIME', '2G话务量',  '2G流量GB', '3G话务量', '3G流量GB', '4G流量GB'])

wbRes_sheetWL = wbRes.create_sheet(index=1,  title='物理基站流量汇总')
wbRes_sheetWL.append(['所属县分', '物理站址名称', '4G流量（GB)', '3G流量（GB)', '3/4G总流量'])

wbRes_sheet4G = wbRes.create_sheet(index=2,  title='4G流量')
wbRes_sheet4G.append(['时间', '物理站点名称', '行政区名称', '小区名称', 'CELL_ID','ENODEB_ID', 'TAC','ENODEB_ID+CELL_ID', '总流量(GB)', '下行PRB平均利用率', 'RRC连接平均数', '平均每PRB干扰噪声功率', '小区级下行单用户平均感知速率', 'RRU型号', '小区下行系统频域带宽','频段指示'])

wbRes_sheet3G = wbRes.create_sheet(index=3,  title='3G话务流量')
wbRes_sheet3G.append(['时间', 'LAC', 'CI', 'LAC+CI', '物理基站名称', '行政区名称', '小区名',	'NodeB名称',	'RNC名称', '话务量', '总流量（GB）', '小区载频平均接收功率[dBm]', '是否3Gonly小区]'])

wbRes_sheet2G = wbRes.create_sheet(index=4,  title='2G话务流量')
wbRes_sheet2G.append(['时间', '物理站点名称', '行政区名称', '小区名称', 'LAC', 'CI', 'LAC+CI', '所属BSC/RNC', 'TCH话务量(erl)', '总流量'])

erl2G = 0 # 2G话务量
total2G = 0 # 2G流量GB
erl3G = 0 # 3G流量GB
total3G = 0 # 3G流量GB
total4G = 0 # 4G流量GB

# 打开2G原始数据
wb2G = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/2G-0729.xlsx')
wb2G_sheet = wb2G.active
for row in range (2, wb2G_sheet.max_row + 1):
    rowStr = str(row)
    colA_2G = wb2G_sheet['A' + rowStr].value  # 时间
    colF_2G = wb2G_sheet['D' + rowStr].value  # CI
    colB_2G = G2R.get(str(colF_2G), "无")  # 物理基站名称
    colC_2G = wb2G_sheet['G' + rowStr].value  # 行政区域名称
    colD_2G = wb2G_sheet['B' + rowStr].value  # 小区名称
    colE_2G = wb2G_sheet['C' + rowStr].value  # LAC
    colG_2G = str(colE_2G) + '-' + str(colF_2G)  # LAC + CI
    colH_2G = wb2G_sheet['E' + rowStr].value  # 所属BSC/RNC
    valueH = wb2G_sheet['H' + rowStr].value  # TCH话务量(erl)
    valueI = wb2G_sheet['I' + rowStr].value
    valueJ = wb2G_sheet['J' + rowStr].value
    if valueH is None:
        valueH = 0
    if valueI is None:
        valueI = 0
    if valueJ is None:
        valueJ = 0

    colJ_2G = round(valueI, 2) + round(valueJ, 2)  # 总流量
    colI_2G = valueH
    erl2G += int(valueH)
    total2G += colJ_2G
    wbRes_sheet2G.append([colA_2G, colB_2G, colC_2G, colD_2G, colE_2G, colF_2G, colG_2G, colH_2G, colI_2G, colJ_2G])
print('完成2G数据')

# 获取3G小区物理基站基础数据,存起来后面会用到
G3R_wb = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/3G小区表.xlsx')
G3R_sheet = G3R_wb.active
G3R = {}
for row in range(2, G3R_sheet.max_row + 1):
   ci = G3R_sheet['E' + str(row)].value
   wljz = G3R_sheet['BY' + str(row)].value
   G3R[ci] = wljz
# print(G3R)

wljzDatas = {}
# 打开3G原始数据
wb3G = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/3G-0729.xlsx')
wb3G_sheet = wb3G.active
for row in range (2, wb3G_sheet.max_row + 1):
    rowStr = str(row)
    colA_3G = wb3G_sheet['A' + rowStr].value  # 时间
    colB_3G = wb3G_sheet['B' + rowStr].value  # LAC
    colC_3G = wb3G_sheet['C' + rowStr].value  # CI
    colD_3G = str(colB_3G) + '-' + str(colC_3G)  # LAC + CI
    colE_3G = G3R.get(str(colC_3G), "无")  # 物理基站名称
    colF_3G = colE_3G[0:2]  # 行政区名，取物理基站前面两位
    colG_3G = wb3G_sheet['D' + rowStr].value  # 小区名称
    colH_3G = wb3G_sheet['E' + rowStr].value  # NodeB名称
    colI_3G = wb3G_sheet['F' + rowStr].value  # RNC名称
    colJ_3G = wb3G_sheet['G' + rowStr].value  # 话务量

    valueH = wb3G_sheet['H' + rowStr].value
    valueI = wb3G_sheet['I' + rowStr].value
    if valueH is None:
        valueH = 0
    if valueI is None:
        valueI = 0
    if colJ_3G is None:
        colJ_3G = 0
    colK_3G = round(valueH, 2) + round(valueI, 2)  # 总流量
    colL_3G = wb3G_sheet['J' + rowStr].value  # 小区载频平均接收功率[dBm]

    erl3G += int(colJ_3G)
    total3G += colK_3G
    wbRes_sheet3G.append([colA_3G, colB_3G, colC_3G, colD_3G, colE_3G, colF_3G, colG_3G, colH_3G, colI_3G, colJ_3G, colK_3G, colL_3G])
    if colE_3G != '无':
        wljzDatas.setdefault(colE_3G, {'3G': colK_3G })


print('完成3G数据')

# 获取4G小区物理基站基础数据,存起来后面会用到
G4R_wb = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/4G小区表.xlsx')
G4R_sheet = G4R_wb.active
G4R = {}
for row in range(2, G4R_sheet.max_row + 1):
   enode = G4R_sheet['I' + str(row)].value
   cell = G4R_sheet['J' + str(row)].value
   wljz = G4R_sheet['BZ' + str(row)].value
   G4R[enode + '-' + cell] = wljz
# print(G4R)

# 打开3G原始数据
wb4G = openpyxl.load_workbook('/Users/wujian/2020项目/话务日报/4G-0729.xlsx')
wb4G_sheet = wb4G.active
for row in range (2, wb4G_sheet.max_row + 1):
    rowStr = str(row)
    colA_4G = wb4G_sheet['A' + rowStr].value  # 时间
    colE_4G = wb4G_sheet['C' + rowStr].value  # CELL_ID
    colF_4G = wb4G_sheet['D' + rowStr].value  # ENODE_ID
    colH_4G = str(colF_4G) + '-' + str(colE_4G)  # ENODE_ID + CELL_ID
    colB_4G = G4R.get(str(colH_4G), "无")  # 物理基站名称
    colC_4G = colB_4G[0:2]  # 行政区名，取物理基站前面两位
    colD_4G = wb4G_sheet['B' + rowStr].value  # 小区名称
    colG_4G = wb4G_sheet['E' + rowStr].value  # TAC
    valueI = wb4G_sheet['I' + rowStr].value
    valueJ = wb4G_sheet['J' + rowStr].value
    if valueI is None:
        valueI = 0
    if valueJ is None:
        valueJ = 0
    colI_4G = round(valueI, 2) + round(valueJ, 2)  # 总流量
    colJ_4G = wb4G_sheet['J' + rowStr].value  # 下行PRB平均利用率
    colK_4G = wb4G_sheet['K' + rowStr].value  # RRC连接平均数
    colL_4G = wb4G_sheet['L' + rowStr].value  # 平均每PRB干扰噪声功率
    colM_4G = wb4G_sheet['M' + rowStr].value  # 小区级下行单用户平均感知速率
    total4G += colI_4G
    wbRes_sheet4G.append([colA_4G, colB_4G, colC_4G, colD_4G, colE_4G, colF_4G, colG_4G, colH_4G, colI_4G, colJ_4G, colK_4G, colL_4G, colM_4G])
    if colB_4G != '无':
        if wljzDatas.get('colB_4G'):
            wljzDatas['colB_4G']['4G'] = colI_4G
        else:
            wljzDatas.setdefault(colB_4G, {'4G': colI_4G})
print('完成4G数据')

# 统计物理基站
print(wljzDatas)
for key in wljzDatas:
    G3L = wljzDatas[key].get('3G', "0")
    G4L = wljzDatas[key].get('4G', "0")
    GTOTAL = round(int(G3L), 2) + round(int(G4L), 2)
    wbRes_sheetWL.append([key[0:2], key, G4L, G3L, GTOTAL])

wbRes_sheet234G.append(['今天', erl2G,  total2G, erl3G, total3G, total4G])

wbRes.save('/Users/wujian/2020项目/话务日报/0729_test.xlsx')
print('完成!!')


